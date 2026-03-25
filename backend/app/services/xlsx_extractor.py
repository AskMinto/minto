"""Deterministic Excel-to-CSV text extractor.

Converts an XLSX/XLS file to a plain-text CSV representation that can be
passed inline to an LLM without needing the Gemini File API.

Strategy (in order):
1. Named tables (openpyxl Table objects) — most structured, preferred.
2. Auto-detected contiguous data blocks per sheet — fallback for Zerodha-style
   sparse layouts where data starts at a variable row with blank rows above.
3. Full sheet dump if no blocks detected.

No LLM involved — pure openpyxl, no network I/O.
"""

from __future__ import annotations

import io
import csv
import logging
from typing import Generator

logger = logging.getLogger(__name__)

# Minimum non-empty cells in a row to count as a data row (not a blank separator)
_MIN_CELLS = 2
# Minimum rows in a block to bother emitting it
_MIN_BLOCK_ROWS = 2


def extract_xlsx(excel_bytes: bytes) -> str:
    """Convert Excel bytes to multi-sheet CSV text.

    Args:
        excel_bytes: Raw (decrypted) XLSX file bytes.

    Returns:
        Plain-text string. Each sheet / table is separated by a blank line.
        Returns empty string on failure.
    """
    try:
        import openpyxl
        wb = openpyxl.load_workbook(
            io.BytesIO(excel_bytes), read_only=False, data_only=True
        )
    except Exception as e:
        logger.error(f"xlsx_extractor: failed to open workbook: {e}")
        return ""

    sections: list[str] = []

    for ws in wb.worksheets:
        sheet_sections = _extract_worksheet(ws)
        sections.extend(sheet_sections)

    return "\n\n".join(sections)


def _extract_worksheet(ws) -> list[str]:
    """Extract all data from a single worksheet, returning a list of CSV sections."""
    sections: list[str] = []
    sheet_name = ws.title

    # ── 1. Named tables ──────────────────────────────────────────────────────
    if hasattr(ws, "tables") and ws.tables:
        for tbl_name, tbl in ws.tables.items():
            rows = list(_iter_table_rows(ws, tbl))
            if rows:
                label = f"# Sheet: {sheet_name} | Table: {tbl_name}"
                sections.append(label + "\n" + _rows_to_csv(rows))
        if sections:
            return sections

    # ── 2. Auto-detect contiguous data blocks ────────────────────────────────
    all_rows = list(ws.iter_rows(values_only=True))
    blocks = list(_detect_blocks(all_rows))

    if blocks:
        for i, block in enumerate(blocks, 1):
            label = f"# Sheet: {sheet_name} | Block {i}"
            sections.append(label + "\n" + _rows_to_csv(block))
        return sections

    # ── 3. Full sheet dump ───────────────────────────────────────────────────
    non_empty = [
        row for row in all_rows
        if sum(1 for v in row if v is not None) >= _MIN_CELLS
    ]
    if non_empty:
        label = f"# Sheet: {sheet_name}"
        sections.append(label + "\n" + _rows_to_csv(non_empty))

    return sections


def _iter_table_rows(ws, tbl) -> Generator[list, None, None]:
    """Yield rows from a named openpyxl Table object."""
    ref = tbl.ref  # e.g. "A1:F50"
    for row in ws[ref]:
        values = [cell.value for cell in row]
        if any(v is not None for v in values):
            yield values


def _detect_blocks(all_rows: list[tuple]) -> Generator[list[list], None, None]:
    """Yield contiguous non-empty row blocks from a worksheet's raw rows.

    A block is a contiguous sequence of rows each having >= _MIN_CELLS non-None
    values. Blocks are separated by rows with fewer than _MIN_CELLS non-None values.
    """
    current_block: list[list] = []

    for row in all_rows:
        non_null_count = sum(1 for v in row if v is not None)
        if non_null_count >= _MIN_CELLS:
            current_block.append(list(row))
        else:
            if len(current_block) >= _MIN_BLOCK_ROWS:
                yield current_block
            current_block = []

    if len(current_block) >= _MIN_BLOCK_ROWS:
        yield current_block


def _rows_to_csv(rows: list[list]) -> str:
    """Serialise a list of rows to a CSV string."""
    buf = io.StringIO()
    writer = csv.writer(buf, lineterminator="\n")
    for row in rows:
        writer.writerow([
            "" if v is None else str(v)
            for v in row
        ])
    return buf.getvalue()
